import io
import json
from typing import Any

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import String, or_
from sqlalchemy.orm import Session

from app.api.deps import current_user, require_admin
from app.models.database import get_db
from app.models.entities import Asset, AssetSegment, User
from app.schemas.common import AssetCreate, AssetImportResult, AssetLookupRequest, AssetOut, AssetUpdate, AssetSegmentCreate, AssetSegmentOut, AssetSegmentUpdate
from app.services.asset_service import (
    build_asset_context,
    find_duplicate_asset,
    lookup_asset_by_domain,
    lookup_asset_by_ip,
    make_asset_key,
    normalize_asset_payload,
)
from app.services.audit_service import write_audit

router = APIRouter(prefix="/assets", tags=["assets"])

FIXED_HEADERS = [
    ("IP", "ip"),
    ("域名", "domain"),
    ("资产名称", "name"),
    ("资产所属区域", "area"),
    ("负责人", "owner"),
    ("部门", "department"),
    ("重要性", "criticality"),
    ("环境", "environment"),
    ("标签", "tags"),
    ("备注", "description"),
]
TEMPLATE_HEADERS = ["IP", "域名", "资产名称", "资产所属区域", "负责人", "部门", "重要性", "环境", "标签", "操作系统", "中间件", "数据库", "开放端口", "业务系统", "备注"]
HEADER_ALIASES = {
    "IP": "ip",
    "ip": "ip",
    "域名": "domain",
    "domain": "domain",
    "资产名称": "name",
    "name": "name",
    "资产所属区域": "area",
    "area": "area",
    "负责人": "owner",
    "owner": "owner",
    "部门": "department",
    "department": "department",
    "重要性": "criticality",
    "criticality": "criticality",
    "环境": "environment",
    "environment": "environment",
    "标签": "tags",
    "tags": "tags",
    "备注": "description",
    "description": "description",
}


@router.get("", response_model=list[AssetOut])
def list_assets(
    q: str | None = None,
    ip: str | None = None,
    domain: str | None = None,
    area: str | None = None,
    owner: str | None = None,
    criticality: str | None = None,
    environment: str | None = None,
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    query = db.query(Asset).filter(Asset.workspace_id == user.workspace_id)
    if ip:
        query = query.filter(Asset.ip.like(f"%{ip.strip()}%"))
    if domain:
        query = query.filter(Asset.domain.like(f"%{domain.strip()}%"))
    if area:
        query = query.filter(Asset.area == area)
    if owner:
        query = query.filter(Asset.owner == owner)
    if criticality:
        query = query.filter(Asset.criticality == criticality)
    if environment:
        query = query.filter(Asset.environment == environment)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(
                Asset.ip.like(like),
                Asset.domain.like(like),
                Asset.name.like(like),
                Asset.owner.like(like),
                Asset.area.like(like),
                Asset.tags.cast(String).like(like),
                Asset.fingerprints.cast(String).like(like),
            )
        )
    return query.order_by(Asset.updated_at.desc()).offset(offset).limit(limit).all()


@router.post("", response_model=AssetOut)
def create_asset(payload: AssetCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    data = normalize_asset_payload(payload.model_dump())
    duplicate = find_duplicate_asset(db, user.workspace_id, data.get("ip", ""), data.get("domain", ""))
    if duplicate:
        raise HTTPException(status_code=409, detail="相同 IP / 域名组合的资产已存在")
    row = Asset(workspace_id=user.workspace_id, asset_key=make_asset_key(user.workspace_id), **data)
    db.add(row)
    db.flush()
    write_audit(db, user, "asset.create", "asset", row.id, {"asset_id": row.id, "ip": row.ip, "domain": row.domain, "name": row.name})
    db.commit()
    db.refresh(row)
    return row


@router.patch("/{asset_id}", response_model=AssetOut)
def update_asset(asset_id: int, payload: AssetUpdate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    row = db.get(Asset, asset_id)
    if not row or row.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="资产不存在")
    
    # 乐观锁检查
    if payload.updated_at and row.updated_at:
        if row.updated_at.replace(microsecond=0) > payload.updated_at.replace(microsecond=0):
            raise HTTPException(status_code=409, detail="资产已被他人修改，请刷新页面后再试")

    original = {
        "ip": row.ip,
        "domain": row.domain,
        "name": row.name,
        "area": row.area,
        "owner": row.owner,
        "department": row.department,
        "criticality": row.criticality,
        "environment": row.environment,
        "tags": row.tags or [],
        "fingerprints": row.fingerprints or {},
        "description": row.description,
    }
    data = normalize_asset_payload({**original, **payload.model_dump(exclude_unset=True, exclude={"updated_at"})})
    duplicate = find_duplicate_asset(db, user.workspace_id, data.get("ip", ""), data.get("domain", ""), exclude_id=row.id)
    if duplicate:
        raise HTTPException(status_code=409, detail="相同 IP / 域名组合的资产已存在")
    changes = {}
    for key, value in data.items():
        old = getattr(row, key)
        if old != value:
            changes[key] = {"old": old, "new": value}
            setattr(row, key, value)
    write_audit(db, user, "asset.update", "asset", row.id, {"asset_id": row.id, "ip": row.ip, "domain": row.domain, "name": row.name, "changes": changes})
    db.commit()
    db.refresh(row)
    return row


@router.delete("/{asset_id}")
def delete_asset(asset_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    row = db.get(Asset, asset_id)
    if not row or row.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="资产不存在")
    detail = {"asset_id": row.id, "ip": row.ip, "domain": row.domain, "name": row.name}
    write_audit(db, user, "asset.delete", "asset", row.id, detail)
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.get("/template.xlsx")
def export_asset_template(db: Session = Depends(get_db), user: User = Depends(current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "资产导入模板"
    ws.append(TEMPLATE_HEADERS)
    ws.append(["10.10.2.15", "portal.example.com", "门户系统", "生产区", "张三", "安全部", "critical", "production", "核心资产,外网暴露", "Ubuntu 22.04", "Nginx", "MySQL", "80,443", "门户业务", "可删除示例行"])
    _add_instruction_sheet(wb, "individual")
    output = _workbook_bytes(wb)
    write_audit(db, user, "asset.export_template", "asset", "template", {"headers": TEMPLATE_HEADERS})
    db.commit()
    return _xlsx_response(output, "asset_template.xlsx")


@router.post("/import", response_model=AssetImportResult)
async def import_assets(
    file: UploadFile = File(...),
    strategy: str = Query("skip", pattern="^(skip|overwrite|append)$"),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [_cell_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if not any(headers):
        raise HTTPException(status_code=400, detail="Excel 第一行必须是字段名")

    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = _row_to_asset(headers, row)
            normalized = normalize_asset_payload(data)
        except Exception as exc:
            stats["skipped"] += 1
            stats["errors"].append({"row": row_index, "error": str(getattr(exc, "detail", exc))})
            continue

        duplicate = find_duplicate_asset(db, user.workspace_id, normalized.get("ip", ""), normalized.get("domain", ""))
        if duplicate and strategy == "skip":
            stats["skipped"] += 1
            continue
        if duplicate and strategy == "overwrite":
            for key, value in normalized.items():
                setattr(duplicate, key, value)
            stats["updated"] += 1
            continue

        db.add(Asset(workspace_id=user.workspace_id, asset_key=make_asset_key(user.workspace_id), **normalized))
        try:
            db.flush()
            stats["created"] += 1
        except Exception:
            db.rollback()
            stats["skipped"] += 1
            stats["errors"].append({"row": row_index, "error": "资产冲突（可能是并发导入了相同 IP/域名）"})
            continue

    write_audit(db, user, "asset.import", "asset", "import", {"filename": file.filename, "strategy": strategy, "stats": stats})
    db.commit()
    return stats


@router.get("/export.xlsx")
def export_assets(
    q: str | None = None,
    area: str | None = None,
    owner: str | None = None,
    criticality: str | None = None,
    environment: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    rows = list_assets(q=q, area=area, owner=owner, criticality=criticality, environment=environment, limit=500, offset=0, db=db, user=user)
    fingerprint_keys = sorted({key for row in rows for key in (row.fingerprints or {}).keys()})
    wb = Workbook()
    ws = wb.active
    ws.title = "资产库"
    fixed_headers = [label for label, _ in FIXED_HEADERS]
    ws.append(fixed_headers + fingerprint_keys)
    for row in rows:
        base = [
            row.ip,
            row.domain,
            row.name,
            row.area,
            row.owner,
            row.department,
            row.criticality,
            row.environment,
            ",".join(row.tags or []),
            row.description,
        ]
        ws.append(base + [row.fingerprints.get(key, "") if row.fingerprints else "" for key in fingerprint_keys])
    output = _workbook_bytes(wb)
    write_audit(db, user, "asset.export", "asset", "export", {"count": len(rows), "fingerprint_columns": fingerprint_keys})
    db.commit()
    return _xlsx_response(output, "assets.xlsx")


@router.post("/lookup")
def lookup_assets(payload: AssetLookupRequest, db: Session = Depends(get_db), user: User = Depends(current_user)):
    by_ip = {}
    by_domain = {}
    for ip in payload.ips:
        ctx = build_asset_context(lookup_asset_by_ip(db, user.workspace_id, ip))
        if ctx:
            by_ip[ip] = ctx
    for domain in payload.domains:
        ctx = build_asset_context(lookup_asset_by_domain(db, user.workspace_id, domain))
        if ctx:
            by_domain[domain] = ctx
    return {"ips": by_ip, "domains": by_domain}


# --- Asset Segments (网段资产) ---

@router.get("/segments", response_model=list[AssetSegmentOut])
def list_segments(
    q: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    query = db.query(AssetSegment).filter(AssetSegment.workspace_id == user.workspace_id)
    if q:
        like = f"%{q.strip()}%"
        query = query.filter(
            or_(
                AssetSegment.segment.like(like),
                AssetSegment.name.like(like),
                AssetSegment.owner.like(like),
                AssetSegment.area.like(like),
            )
        )
    return query.order_by(AssetSegment.updated_at.desc()).all()


@router.post("/segments", response_model=AssetSegmentOut)
def create_segment(payload: AssetSegmentCreate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    # 检查重复
    exists = db.query(AssetSegment).filter_by(workspace_id=user.workspace_id, segment=payload.segment).first()
    if exists:
        raise HTTPException(status_code=409, detail="相同网段已存在")
        
    row = AssetSegment(workspace_id=user.workspace_id, **payload.model_dump())
    db.add(row)
    db.flush()
    write_audit(db, user, "asset_segment.create", "asset_segment", row.id, {"segment": row.segment, "name": row.name})
    db.commit()
    db.refresh(row)
    return row


@router.patch("/segments/{segment_id}", response_model=AssetSegmentOut)
def update_segment(segment_id: int, payload: AssetSegmentUpdate, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    row = db.get(AssetSegment, segment_id)
    if not row or row.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="网段不存在")
    
    # 乐观锁检查
    if payload.updated_at and row.updated_at:
        if row.updated_at.replace(microsecond=0) > payload.updated_at.replace(microsecond=0):
            raise HTTPException(status_code=409, detail="网段已被他人修改，请刷新页面后再试")
        
    data = payload.model_dump(exclude_unset=True, exclude={"updated_at"})
    if "segment" in data:
        duplicate = db.query(AssetSegment).filter(AssetSegment.workspace_id == user.workspace_id, AssetSegment.segment == data["segment"], AssetSegment.id != segment_id).first()
        if duplicate:
            raise HTTPException(status_code=409, detail="相同网段已存在")

    changes = {}
    for key, value in data.items():
        old = getattr(row, key)
        if old != value:
            changes[key] = {"old": old, "new": value}
            setattr(row, key, value)
            
    write_audit(db, user, "asset_segment.update", "asset_segment", row.id, {"segment": row.segment, "changes": changes})
    db.commit()
    db.refresh(row)
    return row


@router.delete("/segments/{segment_id}")
def delete_segment(segment_id: int, db: Session = Depends(get_db), user: User = Depends(require_admin)):
    row = db.get(AssetSegment, segment_id)
    if not row or row.workspace_id != user.workspace_id:
        raise HTTPException(status_code=404, detail="网段不存在")
    write_audit(db, user, "asset_segment.delete", "asset_segment", row.id, {"segment": row.segment})
    db.delete(row)
    db.commit()
    return {"ok": True}


@router.post("/batch-delete")
def batch_delete_assets(payload: dict[str, Any], db: Session = Depends(get_db), user: User = Depends(require_admin)):
    ids = [int(item) for item in payload.get("ids", []) if item]
    if not ids:
        raise HTTPException(status_code=400, detail="请选择资产")
    rows = db.query(Asset).filter(Asset.workspace_id == user.workspace_id, Asset.id.in_(ids)).all()
    count = len(rows)
    for row in rows:
        db.delete(row)
    write_audit(db, user, "asset.batch_delete", "asset", ",".join(str(i) for i in ids), {"count": count})
    db.commit()
    return {"ok": True, "deleted": count}


@router.post("/segments/batch-delete")
def batch_delete_segments(payload: dict[str, Any], db: Session = Depends(get_db), user: User = Depends(require_admin)):
    ids = [int(item) for item in payload.get("ids", []) if item]
    if not ids:
        raise HTTPException(status_code=400, detail="请选择网段")
    rows = db.query(AssetSegment).filter(AssetSegment.workspace_id == user.workspace_id, AssetSegment.id.in_(ids)).all()
    count = len(rows)
    for row in rows:
        db.delete(row)
    write_audit(db, user, "asset_segment.batch_delete", "asset_segment", ",".join(str(i) for i in ids), {"count": count})
    db.commit()
    return {"ok": True, "deleted": count}


# --- Asset Segments (网段资产) Excel Support ---

SEGMENT_HEADERS = ["网段范围", "网段名称", "所属区域", "负责人", "重要性", "环境", "备注说明"]
SEGMENT_HEADER_MAP = {
    "网段范围": "segment",
    "网段名称": "name",
    "所属区域": "area",
    "负责人": "owner",
    "重要性": "criticality",
    "环境": "environment",
    "备注说明": "description",
}

def _add_instruction_sheet(wb: Workbook, type_name: str):
    """
    为模板添加详细的填写说明页签
    """
    ws = wb.create_sheet("填写说明-导入前删除该sheet")
    
    # 样式设置
    from openpyxl.styles import Font, Alignment
    header_font = Font(bold=True, size=12)
    notice_font = Font(bold=True, color="FF0000")
    
    ws.append(["【重要提示：请在开始填写前仔细阅读本页说明】"])
    ws.merge_cells('A1:C1')
    ws['A1'].font = notice_font

    ws.append([]) # 空行
    ws.append(["字段名称", "是否必填", "填写要求与示例"])
    
    # 设置表头样式
    for cell in ws[3]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    if type_name == "individual":
        guidelines = [
            ["IP", "二选一", "主机的 IPv4 或 IPv6 地址。"],
            ["域名", "二选一", "主机的完整域名 (FQDN)。IP 和域名至少填写一个，用于日志自动关联。"],
            ["资产名称", "是", "业务可读名称，如：财务系统主库、外部门户 Web 服务器。"],
            ["资产所属区域", "否", "资产所在的逻辑或物理区域，如：DMZ区、生产区、办公区、管理区。"],
            ["负责人", "否", "该资产的第一责任人姓名。"],
            ["重要性", "否", "可选值：low(低), medium(中), high(高), critical(极高)。默认为 medium。"],
            ["环境", "否", "部署环境，推荐：生产、测试、开发、办公、隔离区。"],
            ["标签", "否", "多个标签请用英文逗号分隔，如：外网暴露,重要业务,核心交换。"],
            ["备注", "否", "关于该资产的其他补充描述。"],
            ["---", "---", "---"],
            ["【高级指纹逻辑】", "核心功能", "除上述固定列外，您可以在 Excel 中【自由增加任意数量的列】。"],
            ["", "", "这些额外列的【表头】将自动识别为指纹名，【内容】识别为指纹值。"],
            ["", "", "示例：您可以增加列“操作系统”、“数据库版本”、“开放端口”，系统会全自动提取。"],
        ]
    else:
        guidelines = [
            ["网段范围", "是", "支持 CIDR 格式 (如 192.168.1.0/24) 或范围格式 (如 10.0.0.1-50)。"],
            ["网段名称", "是", "描述该网段的业务用途，如：呼叫中心办公网段。"],
            ["所属区域", "否", "该网段对应的逻辑区域，如：DMZ区、生产区。"],
            ["负责人", "否", "该网段的运维或安全责任人。"],
            ["重要性", "否", "可选值：low(低), medium(中), high(高), critical(极高)。默认为 medium。"],
            ["环境", "否", "建议填：生产、测试、开发、办公、隔离区。"],
            ["备注说明", "否", "对该网段用途的详细说明。"],
        ]

    for row in guidelines:
        ws.append(row)

    # 调整列宽
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 100
    
    # 将“填写说明”页签置于首位
    wb._sheets = [wb._sheets[-1]] + wb._sheets[:-1]


@router.get("/segments/template.xlsx")
def export_segment_template(db: Session = Depends(get_db), user: User = Depends(current_user)):
    wb = Workbook()
    ws = wb.active
    ws.title = "网段资产导入模板"
    ws.append(SEGMENT_HEADERS)
    ws.append(["192.168.1.0/24", "测试网段", "测试区", "张三", "medium", "test", "示例行"])
    _add_instruction_sheet(wb, "segment")
    output = _workbook_bytes(wb)
    return _xlsx_response(output, "asset_segment_template.xlsx")


@router.get("/segments/export.xlsx")
def export_segments_xlsx(
    q: str | None = None,
    db: Session = Depends(get_db),
    user: User = Depends(current_user),
):
    rows = list_segments(q=q, db=db, user=user)
    wb = Workbook()
    ws = wb.active
    ws.title = "网段资产库"
    ws.append(SEGMENT_HEADERS)
    for row in rows:
        ws.append([
            row.segment,
            row.name,
            row.area,
            row.owner,
            row.criticality,
            row.environment,
            row.description,
        ])
    output = _workbook_bytes(wb)
    return _xlsx_response(output, "asset_segments.xlsx")


@router.post("/segments/import", response_model=AssetImportResult)
async def import_segments(
    file: UploadFile = File(...),
    strategy: str = Query("skip", pattern="^(skip|overwrite|append)$"),
    db: Session = Depends(get_db),
    user: User = Depends(require_admin),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    headers = [_cell_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    
    stats = {"created": 0, "updated": 0, "skipped": 0, "errors": []}
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {}
            for header, val in zip(headers, row):
                key = SEGMENT_HEADER_MAP.get(header)
                if key:
                    data[key] = _cell_text(val)
            
            if not data.get("segment"):
                continue
                
            duplicate = db.query(AssetSegment).filter_by(workspace_id=user.workspace_id, segment=data["segment"]).first()
            if duplicate and strategy == "skip":
                stats["skipped"] += 1
                continue
            if duplicate and strategy == "overwrite":
                for k, v in data.items():
                    setattr(duplicate, k, v)
                stats["updated"] += 1
                continue
            
            db.add(AssetSegment(workspace_id=user.workspace_id, **data))
            try:
                db.flush()
                stats["created"] += 1
            except Exception:
                db.rollback()
                stats["skipped"] += 1
                stats["errors"].append({"row": row_index, "error": "网段冲突（可能是并发导入了相同网段）"})
                continue
        except Exception as exc:
            stats["skipped"] += 1
            stats["errors"].append({"row": row_index, "error": str(exc)})
            
    db.commit()
    return stats


def _row_to_asset(headers: list[str], row: tuple[Any, ...]) -> dict[str, Any]:
    data: dict[str, Any] = {"fingerprints": {}}
    for header, raw_value in zip(headers, row):
        if not header:
            continue
        value = _cell_text(raw_value)
        key = HEADER_ALIASES.get(header.strip())
        if key == "tags":
            data["tags"] = [item.strip() for item in value.replace("，", ",").split(",") if item.strip()]
        elif key:
            data[key] = value
        elif value:
            data["fingerprints"][header.strip()] = value
    return data


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value).strip()


def _workbook_bytes(wb: Workbook) -> bytes:
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _xlsx_response(content: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        iter([content]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
